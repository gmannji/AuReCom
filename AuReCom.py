# auto review risk register
#!/usr/bin/env python
# -*- coding: utf-8 -*-
#  
#  Copyright 2017 hazmanyusoff 
#  

# import packages
import openpyxl # main module 1
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side # for styling - self explanatory
from openpyxl.utils import get_column_letter # to convert numbers into columns alphabets in excel
import pandas as pd # main module 2
import time # to get time
from datetime import datetime, date, timedelta # to get datetime format and adding days
import csv # to read csv file
import os.path # to get file path name
import collections # to count number of occurences in a list
from tkinter.filedialog import askopenfilename # open dialog to select file
from tkinter import Tk # to remove background box when select file
import re # to use regular expression
from itertools import zip_longest # to get 2 lists element-wise into tuples of a list

# Split by line function for simplicity
def splitline (abc):
    return abc.split('\n')

# Split by dash function for simplicity
def splitdash (xyz):
    return xyz.split(' - ')

#class color:
   #PURPLE = '\033[95m'
   #CYAN = '\033[96m'
   #DARKCYAN = '\033[36m'
   #BLUE = '\033[94m'
   #GREEN = '\033[92m'
   #YELLOW = '\033[93m'
   #RED = '\033[91m'
   #BOLD = '\033[1m'
   #UNDERLINE = '\033[4m'
   #END = '\033[0m'

# getting the date today and 90 days from now
today = date.today()
three_months = today + timedelta(days=90)
thirty_days = today + timedelta(days=30)

# Cell format
redFill = PatternFill(start_color='FFC7CE',
                      end_color='FFC7CE',
                      fill_type='solid')
yellowFill = PatternFill(start_color='FFFFCC', end_color='FFFFCC',
                         fill_type='solid')
greyFill = PatternFill(start_color='DCDCDC',
                      end_color='DCDCDC',
                      fill_type='solid')
blueFill = PatternFill(start_color='B8CCE4',
                      end_color='B8CCE4',
                      fill_type='solid')
tanFill = PatternFill(start_color='C4BD97',
                      end_color='C4BD97',
                      fill_type='solid')
greenFill = PatternFill(start_color='C6EFCE',
                      end_color='C6EFCE',
                      fill_type='solid')
orangeFill = PatternFill(start_color='FFEB9C',
                      end_color='FFEB9C',
                      fill_type='solid')
italic = Font(size=10, italic=True)
alignment = Alignment(horizontal='general',
                      vertical='top',
                      text_rotation=0,
                      wrap_text=True,
                      shrink_to_fit=False,
                      indent=0)
border = Border(left=Side(border_style='thin',
                          color='B2B2B2'),
                right=Side(border_style='thin',
                           color='B2B2B2'),
                top=Side(border_style='thin',
                         color='B2B2B2'),
                bottom=Side(border_style='thin',
                            color='B2B2B2'))

# Hide Tk window pop up when selecting file
root = Tk()
root.withdraw()

# Select file
print('Please select a risk register to review...\n NOTE: Please ensure the risk register is\n in 03. RAPID Risk Register format from \n ARM Custom Reports\n')
filenew = askopenfilename()
if filenew is '':
    print('No file selected\n\nExiting...\n\n\nDeveloped by Hazman Yusoff')
    time.sleep(2)
    quit()

# file path as per selected file. this is the new file will be saved in the same directory
file_path = os.path.splitext(filenew)[0]+'-comments.xlsx'
print('Reviewing...')

# Load workbook and sheet as per selected file
wb = openpyxl.load_workbook(filenew)
sheet = wb.get_sheet_by_name('03. RAPID Risk Register')

# comments row
sheet['X3'].fill = orangeFill
sheet['Y3'] = 'Attention'
sheet['X4'].fill = redFill
sheet['Y4'] = 'To improve. Refer comments at the top or refer to SRE'
sheet['AD3'].fill = blueFill
sheet['AE3'] = 'To re-assess - Actions are completed'
sheet['AD4'].fill = tanFill
sheet['AE4'] = 'To re-assess - Action(s) still Active'
sheet['AJ3'].fill = greenFill
sheet['AK3'] = "What's changed"
sheet.row_dimensions[5].height = 130
sheet['B5'] = 'Comments'
sheet['D5'] = "Consider to re-title risk that starts with 'Delay…'.\nRefer to ARM App User guide in step 2 page 19 and update where required."
sheet['E5'] = 'Risk Owner will always be the package head / manager.\nRefer to ARM App User guide in step 2 page 20 and update where required.'
sheet['G5'] = 'Please review the unapproved risks if any.\nRefer to ARM App User guide in step 2 page 21 - 23 and update where required.'
sheet['H5'] = 'Refer to ARM App User guide in step 2 page 24 and update where required'
sheet['I5'] = 'Refer to ARM App User guide in step 2 page 25 and update where required'
sheet['K5'] = 'Please ensure the current controls are cross referenced to the causes.\nRefer to ARM App User guide in step 2 page 26 - 27 and update where required'
sheet['L5'] = 'High risks require more description and info.\nRefer to ARM App User guide in step 2 page 28 - 29 and update where required'
sheet['M5'] = 'Which activity ID does this risk impact on?\nRefer to ARM App User guide in step 2 page 30 and update'
sheet['N5'] = 'Refer to ARM App User guide in step 2 page 32-33 and update where required'
sheet['O5'] = 'Refer to ARM App User guide in step 2 page 34 and update where required'
sheet['U5'] = 'Review and update providing a brief concise note of how you arrived at the assessment\n"Probability of X% based on …"\n"Impact of X-Y days to recover"\nRefer to ARM App User guide in step 2 page 62 and update.'
sheet['V5'] = 'Note the dates past the resolution date. Have these been resolved? If so has the risk been updated or if not what is happening?\nRefer to ARM App User guide in step 2 page 49 and update.'
sheet['W5'] = 'Please note these dates should be aligned to your schedule activity that you have noted in the activity field where applicable.\nRefer to ARM App User guide in step 2 page 50 and update where required.'
sheet['X5'] = 'Please note these dates should be aligned to your schedule activity that you have noted in the activity field where applicable.\nRefer to ARM App User guide in step 2 page 51 and update where required.'
sheet['AA5'] = 'At the moment, we are not assessing other impact than Schedule'
sheet['AB5'] = 'At the moment, we are not assessing other impact than Schedule'
sheet['AC5'] = 'At the moment, we are not assessing other impact than Schedule'
sheet['AD5'] = 'Current risk rating that is lower than target risk rating shall be re-assessed.\nAny high risk to be reviewed and approved?'
sheet['AG5'] = 'At the moment, we are not assessing other impact than Schedule'
sheet['AH5'] = 'At the moment, we are not assessing other impact than Schedule'
sheet['AI5'] = 'At the moment, we are not assessing other impact than Schedule'
sheet['AK5'] = 'Are there any additional actions?\nPlease ensure the actions are cross referenced to the causes.\nRefer to ARM App User guide in step 2 page 79-80 and update where required.'
sheet['AL5'] = "Review 'Undefined' action owners for Active and Completed actions. Marked with '!!!'\nHave the action owners changed for any outstanding actions?\nRefer to ARM App User guide in step 2 page 81 and update where required."
sheet['AM5'] = "Have all completed actions been updated and if completed then has the current risk rating been re-assessed?\nHave all proposed actions been reviewed and made active? Proposed actions are marked with '!!!'\nIf there is any active action and current and target risk rating the same, please re-assess.\nRefer to ARM App User guide in step 2 page 82 and update where required."
sheet['AN5'] = 'Refer to ARM App User guide in step 2 page 83 and update where required.'
sheet['AR5'] = "Are there any active actions with no due date?\nAre there any overdue actions?\nOverdue actions are marked with '!!!'\nActions marked with '@' are due within 30 days"

# Getting the first row and all column name
# Read excel with no header to get accurate row number and fill nan with 'N/A' for easy tracking as our registers have merged columns and giving blank columns
new = pd.read_excel(filenew,header=None).fillna('N/A')

# iterate each row into tuples then into list
cycle=list(new.itertuples(index=False))

# empty list for our column name. this will be total unmodified list
columnlist_ori=[]

# loop rows in cycle
header_df=0 # pandas starts with 0. openpyxl starts with 1
while header_df < len(cycle):
    for colu in cycle[header_df]:
        columnlist_ori.append(colu)
    counter=collections.Counter(columnlist_ori) # Counting occurences in the columnlist
    if counter['N/A'] < 5: # set the value for blank columns less than 5. this should give good impression that the row is the first row as header.
        break
    header_df+=1
    del columnlist_ori[:] # clear columnlist as we are using append - it won't overwrite.

# define headers and firstrow
SRE_comment_row = header_df
header_excel = header_df + 1
firstrow_excel = header_excel + 1

# Calculate no of risks and no of ratings
# re-define 'new' df
new = pd.read_excel(filenew, header=header_df)
#print(new['Status'].value_counts())
freq = new['Status'].value_counts() # get the frequency 'table'
#print(freq['Open - Active'])

# no of ratings
# filter Open-Active and Rating into variables
very_high_open =  (new['Status'] ==  'Open - Active') & (new['Current Risk Level'].str.endswith('Very High'))
high_open =  (new['Status'] ==  'Open - Active') & (new['Current Risk Level'].str.endswith(': High'))
medium_open =  (new['Status'] ==  'Open - Active') & (new['Current Risk Level'].str.endswith('Medium'))
low_open =  (new['Status'] ==  'Open - Active') & (new['Current Risk Level'].str.endswith('Low'))
nil_open = (new['Status'] ==  'Open - Active') & (new['Current Risk Level'].str.endswith('NIL'))

# get the number
no_very_high_open = new[very_high_open].shape[0]
no_high_open =  new[high_open].shape[0]
no_medium_open =  new[medium_open].shape[0]
no_low_open =  new[low_open].shape[0]
no_nil_open = new[nil_open].shape[0]
RATING_SUMMARY =[] # define list to be appended later

if no_very_high_open != 0:
    RATING_SUMMARY.append(str(no_very_high_open) + ' Very High Risk(s)')
if no_high_open != 0:
    RATING_SUMMARY.append(str(no_high_open) + ' High Risk(s)')
if no_medium_open != 0:
    RATING_SUMMARY.append(str(no_medium_open) + ' Medium Risk(s)')
if no_low_open != 0:
    RATING_SUMMARY.append(str(no_low_open) + ' Low Risk(s)')
if no_nil_open != 0:
    RATING_SUMMARY.append(str(no_nil_open) + ' Risk(s) Not Rated')

# no of risks
# initialize no of risks
no_open_risks = 0
no_unapproved_risks = 0
no_closed_risks = 0
RISK_COUNT = []

# Count no of risks. Use try because there may be no unapproved risks etc.
try:
    no_open_risks = freq['Open - Active']
except KeyError:
	pass

try:
    no_unapproved_risks = freq['Unapproved']
except KeyError:
	pass

try:
    no_closed_risks = freq['Closed']
except KeyError:
	pass

if no_open_risks != 0:
    RISK_COUNT.append(str(no_open_risks) + ' Open - Active Risk(s)')
    #print(RISK_COUNT)
if no_unapproved_risks != 0:
    RISK_COUNT.append(str(no_unapproved_risks) + ' Unapproved Risk(s)')
    #print(RISK_COUNT)
if no_closed_risks != 0:
    RISK_COUNT.append(str(no_closed_risks) + ' Closed Risk(s)')
    #print(RISK_COUNT)

#print (no_open_risks)
#print(no_unapproved_risks)
#print(no_closed_risks)

#no_open_risks = 'No. of Open - Active risks: ' + str(freq['Open - Active'])

#try:
	##print(freq['Unapproved'])
	#count_unapproved_risks = 'No. of Unapproved risks: ' + str(freq['Unapproved'])
#except KeyError:
	#pass

# all lists define here
columnlist = [x for x in columnlist_ori if x != 'N/A'] # clear out the 'N/A'
COLUMN_NO = dict((keys,int) for keys in columnlist) # getting the columns into dict as keys. this is so we can assign column numbers as its values. we can also use columnlist.index('Column Name')+1 but this is more intuitive
UPDATE = ['Undefined', 'Unapproved', '0 : NIL']
# counting actions
OVERDUE = 0
ACTIVE = 0
PROPOSED = 0
COMPLETE = 0
ABANDON = 0
ACTION_DUE = 0
ACTIONS_SUMMARY=[]
TARGET_RES = []
TRIGGERING = []
EXPIRING = []
EXPIRED = []
TRIG_DATES = []
# regex for current controls and actions
valid = ['(all)','note:']
c1c2c3 = re.compile(r'c\d') #regex created

# looping the sheet
for rowNum in range(SRE_comment_row, sheet.max_row + 1):
    for columnNum in range(1, sheet.max_column + 1):

        # setting up variables for loop
        cell = sheet.cell(row=rowNum, column=columnNum)

        # settign the format for comments row
        if rowNum == SRE_comment_row and cell.value is not None:
            cell.font = italic
            cell.fill = yellowFill
            cell.alignment = alignment
            cell.border = border

        # getting column number into dictionary value
        elif rowNum == header_excel and cell.value in COLUMN_NO:
            COLUMN_NO[cell.value] = columnNum

        # generic cell condition that only need redFill
                          # cell is blank or in UPDATE. Skip few columns as not used.
        elif rowNum >= firstrow_excel and (((columnNum != COLUMN_NO['Description'] and columnNum != COLUMN_NO['Reviewed By'] and columnNum != COLUMN_NO['Last Review'] and columnNum != COLUMN_NO['Next Review Date'] and columnNum != COLUMN_NO['Last Review Note']) and
                          (cell.value == '' or cell.value is None or cell.value in UPDATE)) or 
                          # risk title that starts with delay
                          (columnNum == COLUMN_NO['Title'] and cell.value.startswith("Delay")) or
                          # risk rating rated other than schedule impact
                          ((columnNum == COLUMN_NO['Current Cost Impact'] or
                                   columnNum == COLUMN_NO['Current Reputation Impact '] or
                                   columnNum == COLUMN_NO['Current HSE Impact'] or
                                   columnNum == COLUMN_NO['Target Cost Impact'] or
                                   columnNum == COLUMN_NO['Target Reputation Impact'] or
                                   columnNum == COLUMN_NO['Target HSE Impact']) and cell.value != 'NIL')
                          ):
            cell.fill = redFill
        
        # description, reviewed by, last review, next review date, last review note columns to be filled in if high risk
        elif rowNum >= firstrow_excel and (columnNum == COLUMN_NO['Description'] or columnNum == COLUMN_NO['Reviewed By'] or columnNum == COLUMN_NO['Last Review'] or columnNum == COLUMN_NO['Next Review Date'] or columnNum == COLUMN_NO['Last Review Note']):
            currate = sheet.cell(row=rowNum, column=COLUMN_NO['Current Risk Level'])
            tarrate = sheet.cell(row=rowNum, column=COLUMN_NO['Target Risk Level'])
            curratelist = currate.value.split(' : ')
            if int(curratelist[0]) >= 16 and (cell.value == '' or cell.value is None):
                cell.fill = redFill

        # current controls and action title cross reference
        elif rowNum >= firstrow_excel and (columnNum == COLUMN_NO['Current Controls'] or columnNum == COLUMN_NO['Action Title']):
            cxreflist = splitline(cell.value)
            cxreflist = [x for x in cxreflist if x.replace(" ", "") != ''] # deleting blank lines and blank lines that have spaces
            for cx in cxreflist:
                action = cx.lower().replace(" ", "") # standardize format for easiness of searching
                cxsearch = c1c2c3.search(action)

                # if c1% or valid words found in the string or action is blank if risk is low, continue
                if cxsearch is not None or any(word in action for word in valid) or (action == '-' and int(curratelist[0]) < 11):
                    continue
                else:
                    cell.fill = redFill
                    break

        # target resolution
        elif rowNum >= firstrow_excel and columnNum == COLUMN_NO['Target Resolution']:
            riskID = sheet.cell(row=rowNum, column=COLUMN_NO['ID'])
            res_date = cell.value.date()
            if res_date < today and currate.value != tarrate.value:
                TARGET_RES.append(riskID.value)
                cell.fill = redFill

        # trigger date
        elif rowNum >= firstrow_excel and columnNum == COLUMN_NO['Trigger Date']:
            #trig_date = datetime.strptime(cell.value, '%d %b %y').date()
            #trig_date= datetime.datetime.now().date()
            riskID = sheet.cell(row=rowNum, column=COLUMN_NO['ID'])
            trig_date = cell.value.date()
            if trig_date > today and trig_date < three_months:
                TRIGGERING.append(riskID.value)
                cell.fill = orangeFill
                #print('risk to be triggered - ' + riskID.value)
                
        # expiry date        
        elif rowNum >= firstrow_excel and columnNum == COLUMN_NO['Expiry Date']:
            exp_date = cell.value.date()
            if exp_date > today and exp_date < three_months:
                EXPIRING.append(riskID.value)
                cell.fill = orangeFill
                #print('risk expiring soon - ' + riskID.value)
            elif exp_date < today:
                EXPIRED.append(riskID.value)
                cell.fill = redFill
                #print('risk expired - ' + riskID.value)

        # current rating lesser than target rating
        elif rowNum >= firstrow_excel and columnNum == COLUMN_NO['Current Risk Level']:
            tarratelist = tarrate.value.split(' : ')  # split the numbers and word
            if int(curratelist[0]) < int(tarratelist[0]):  # compare the number rating
                cell.fill = redFill

        # treatment actions
        elif rowNum >= firstrow_excel and columnNum == COLUMN_NO['Action Status']:
            stslist = splitline(cell.value) # getting each line of action status into list

            riskstatus = sheet.cell(row=rowNum, column=COLUMN_NO['Status'])
            duedates = sheet.cell(row=rowNum, column=COLUMN_NO['Due Date'])
            duedatelist = splitline(duedates.value) # getting each line of action duedate into list

            actowner = sheet.cell(row=rowNum, column=COLUMN_NO['Action Owner'])
            actownerlist = splitline(actowner.value) # getting each line of action owner into list
            
            # loop for action status by line
            for idx, val in enumerate(stslist):
                if val.endswith("Active"):
                    if riskstatus.value == 'Open - Active':
                        ACTIVE += 1 # number of active actions
                    date_line = splitdash(duedatelist[idx]) # splitting the action id and date into list. d[0] is the action id. d[1] is the date to compare.
                                        
                    if date_line[1] != '':  # if the date is not empty
                        due = datetime.strptime(date_line[1], '%d %b %y').date()  # converting str type into datetime type
                        if due < today:  # compare with today's date
                            duedatelist[idx] = duedatelist[idx] + " !!!" #this will just update the corresponding date in list
                            if riskstatus.value == 'Open - Active':
                                OVERDUE += 1
                            duedates.fill = redFill
                        elif due > today and due < thirty_days: # due in 30 days
                            if riskstatus.value == 'Open - Active':
                                ACTION_DUE += 1
                            duedatelist[idx] = duedatelist[idx] + " @"
                            duedates.fill = orangeFill

                    else:  # if date is empty
                        duedates.fill = redFill

                    if actownerlist[idx].endswith('Undefined'): # if action owner is undefined for active actions
                        actownerlist[idx] = actownerlist[idx] + " !!!"
                        actowner.fill = redFill                    

                elif val.endswith('Proposed'):
                    if riskstatus.value == 'Open - Active':
                        PROPOSED += 1 # number of proposed actions
                    stslist[idx] = stslist [idx] + " !!!"
                elif val.endswith('Complete'):
                    if riskstatus.value == 'Open - Active':
                        COMPLETE += 1 # number of completed actions
                    if actownerlist[idx].endswith('Undefined'): # if action owner is undefined for completed actions
                        actownerlist[idx] = actownerlist[idx] + " !!!"
                        actowner.fill = redFill
                elif val.endswith('Abandon'):
                    if riskstatus.value == 'Open - Active':
                        ABANDON += 1 # number of abandoned actions
            
            cell.value = '\n'.join(stslist) # join as we have modify proposed actions by adding '!!!'
            duedates.value = '\n'.join(duedatelist) # join all dates into the due date cell as have modify due date by adding '!!!'
            actowner.value = '\n'.join(actownerlist) # join as we have modify action owner by adding '!!!'
            #print(str(test) + 'actions due in 30 days')

            ''' completed action but not re-assessed, 
                actions are still active but rating the same,
                proposed actions'''
            if cell.value.find("Active") == -1 and currate.value != tarrate.value:# or (cell.value.find("Active") != -1 and currate.value == tarrate.value):
                cell.fill = blueFill
                currate.fill = blueFill
                tarrate.fill = blueFill
            elif cell.value.find("Active") != -1 and currate.value == tarrate.value:
                cell.fill = tanFill
                currate.fill = tanFill
                tarrate.fill = tanFill                
            if  cell.value.find('!!!') != -1:
                cell.fill = redFill

# count overdue active actions percentage
try:
    percent = round((OVERDUE/ACTIVE*100),1)
except ZeroDivisionError:
    percent = 0

# append and display trigger dates and use join to join riskID
if TARGET_RES:
    TRIG_DATES.append(str(len(TARGET_RES)) + ' Targeted Risk(s) Unresolved')
if TRIGGERING:
    TRIG_DATES.append(str(len(TRIGGERING)) + ' Risk(s) Triggering within 90 days')# + ' (' + ', '.join(TRIGGERING) + ')')
if EXPIRING:
    TRIG_DATES.append(str(len(EXPIRING)) + ' Risk(s) Expiring within 90 days')# + ' (' + ', '.join(EXPIRING) + ')')
if EXPIRED:
    TRIG_DATES.append(str(len(EXPIRED)) + ' Risk(s) Expired')# + ' (' + ', '.join(EXPIRED) + ')')

# append and display all actions by numbers
ACTIONS_SUMMARY.append(str(OVERDUE) + ' Overdue Action(s). ' + str(percent) + ' % of actions are overdue')
ACTIONS_SUMMARY.append(str(ACTIVE) + ' Active Action(s). ' + str(ACTION_DUE) + ' Due within 30 days.')
ACTIONS_SUMMARY.append(str(PROPOSED) + ' Proposed Action(s).')
ACTIONS_SUMMARY.append(str(COMPLETE) + ' Completed Action(s).')
ACTIONS_SUMMARY.append(str(ABANDON) + ' Abandoned Action(s).')
# join all RISK_COUNT
RISK_COUNT = '\n'.join(RISK_COUNT)
# join all TRIG_DATES
TRIG_DATES = '\n'.join(TRIG_DATES)
# join all RATING_SUMMARY
RATING_SUMMARY = '\n'.join(RATING_SUMMARY)
# join all ACTIONS_SUMMARY
ACTIONS_SUMMARY = '\n'.join(ACTIONS_SUMMARY)
# setting parameters and format for the comment summary cell
cell_final = sheet.cell(row=firstrow_excel, column=sheet.max_column + 2)
cell_final.alignment = alignment
cell_final.border = border
cell_final.fill = yellowFill
# setting cell summary (1 column extra to rightmost cell) width and unhide if hidden
sheet.column_dimensions[get_column_letter(len(columnlist_ori)+2)].width = 50 # get_column_letter input is int, output alphabet i.e column letter
sheet.column_dimensions[get_column_letter(len(columnlist_ori)+2)].hidden = False

FINAL_SUMMARY = 'Risk Status Summary:' + '\n' + RISK_COUNT + '\n' + TRIG_DATES + '\n\n' + 'Risk Rating Summary (Open-Active Risks):' + '\n' + RATING_SUMMARY + '\n\n' + 'Action Summary (Open-Active Risks):' + '\n' + ACTIONS_SUMMARY
cell_final.value = FINAL_SUMMARY

# re-title sheet and save
sheet.title='Comments'
wb.save(file_path)

print('\nDone!\n\nNow select a risk register with the same\nformat to make a comparison\n')

# ask for 2nd risk register to do comparison
fileold = askopenfilename()
if fileold is '':
    print('No file selected\n\nExiting...\n\n\nDeveloped by Hazman Yusoff')
    time.sleep(3)
    quit()
else:
    pass

print('Comparing...\n')


# Define the diff function to show the changes between 2 risk registers
def report_diff(x):
    if x[0] != x[1]:# and (x[0] is not None or x[0] != 'nan'):
        return ('{}\n--<<>>--\n{}'.format(*x))
    else:
        return x[0]
  #return x[0] if x[0] == x[1] else '{}\n--<<>>--\n{}'.format(*x)

def sorted_nicely( l ): 
    """ Sort the given iterable in the way that humans expect.""" 
    convert = lambda text: int(text) if text.isdigit() else text 
    alphanum_key = lambda key: [ convert(c) for c in re.split('([0-9]+)', key) ] 
    return sorted(l, key = alphanum_key)

#def sort_human(l):
    #convert = lambda text: float(text) if text.isdigit() else text
    #alphanum = lambda key: [ convert(c) for c in re.split('([-+]?[0-9]*\.?[0-9]*)', key) ]
    #l.sort( key=alphanum )
    #return l

def compare_sort (cellvalue):
    cell_old,cell_new = cellvalue.split('\n--<<>>--\n') # split what's old and new into 2 strings
    cell_old = cell_old.split('\n') # update the split list and split further based on line items 
    cell_new = cell_new.split('\n') # update the split list and split further based on line items 
    
    # call sorted nicely to sort alphanumerically
    for x in sorted_nicely(cell_old):
        cell_old_sorted.append(x)    
    for x in sorted_nicely(cell_new):
        cell_new_sorted.append(x)
    
    # combine the sorted into a list of tuple and return
    cell_both = list(zip_longest(cell_old_sorted,cell_new_sorted))
    return cell_both

# ignore warnings when comparing. just a warning using old method
# warnings.filterwarnings("ignore")

# disable SettingWithCopyWarning 
pd.options.mode.chained_assignment = None  # default='warn'

# Read in the two files but call the data old and new and create columns to track. dropna not used anymore as some columns we want may have na values
old = pd.read_excel(fileold, header=header_df)
#new = pd.read_excel(filenew, header=header_df)
old['version'] = "old"
new['version'] = "new"

# empty dict to get column width values
colwidth = {}

# Join all the data together and ignore indexes so it all gets added
# dropna here and columns order is still intact. initially used to drop merged columns or blank columns. not using anymore...as this will dropna for blank column but we want that column(i.e reviewed by - most of time will be blank)
full_set = pd.concat([old,new],ignore_index=True).fillna('')

# Let's see what changes in the main columns we care about. Whatever the same, it'll get removed - subset is crucial if not it won't remove duplicates.
changes = full_set.drop_duplicates(subset=columnlist,keep='last')

# Now the duplicates are removed, we want to know where the duplicate RiskIDs are, that means there have been changes. this return only the duplicated RiskID
dupe_riskID = changes.set_index("ID").index.get_duplicates()

# Get all the duplicate rows that are in dupe_riskID. Subset into new dataframe
dupes = changes[changes["ID"].isin(dupe_riskID)]

# Pull out the old and new data into separate dataframes
change_new = dupes[(dupes["version"] == "new")]
change_old = dupes[(dupes["version"] == "old")]

# Drop the temp columns - we don't need them now
change_new = change_new.drop(['version'], axis=1)
change_old = change_old.drop(['version'], axis=1)

# Index on the RiskID
change_new.set_index('ID',inplace=True)
change_old.set_index('ID',inplace=True)

# Now we can diff because we have two data sets of the same size with the same index
# check to see if change is empty. meaning no changes at all. if we dont check, empty dataframe will give error.
if change_new.empty == True and change_old.empty == True:
    changed_risks = change_new
    
else:
    diff_panel = pd.Panel(dict(df1=change_old,df2=change_new))
    changed_risks = diff_panel.apply(report_diff, axis=0)
    #changed_risks.ActionStatus.str.contains("<<>>")

# Diff'ing is done, we need to get a list of removed risks

# Flag all duplicated RiskIDs - risks that have changed.
changes['duplicate']=changes["ID"].isin(dupe_riskID)

# Identify non-duplicated risks that are in the old version and did not show in the new version
closed_risks = changes[(changes["duplicate"] == False) & (changes["version"] == "old")]

# Re-Index Closed Risks
closed_risks.set_index('ID',inplace=True)

# We have closed risks and changed risks, we need to figure out which ones are new

# Drop duplicates but keep the first item instead of the last
new_risk_set = full_set.drop_duplicates(subset=columnlist,keep='first')

# Identify dupes in this new dataframe
new_risk_set['duplicate']=new_risk_set["ID"].isin(dupe_riskID)

# Identify added risks
added_risks = new_risk_set[(new_risk_set["duplicate"] == False) & (new_risk_set["version"] == "new")]

# New risks into set so that we can highlight risk ID cells for new risks
added_risks_set = set(added_risks['ID'])

# Re-Index added risks
added_risks.set_index('ID',inplace=True)

# Save the changes to excel but only include the columns we care about. use openpyxl engine to add new sheets. else it'll overwrite
book=openpyxl.load_workbook(file_path)
writer = pd.ExcelWriter(file_path,engine='openpyxl')
writer.book=book

# Write on top of active sheet - no overwrite - no adding new sheets
writer.sheets=dict((ws.title,ws) for ws in book.worksheets)

# to remove the index column (RiskID) from the columnlist as we don't need it for new sheets. else it'll get duplicated due to indexing in pandas.
columnlist_ID_removed = list(columnlist)
columnlist_ID_removed.remove(changed_risks.index.name)

# save to an existing file in a new sheet and select the columns we want
changed_risks.to_excel(writer,"Changed", index=True, columns=columnlist_ID_removed + ['Actions Tracker']) # added extra column 'Actions Tracker' for tracking
closed_risks.to_excel(writer,"Closed",index=True, columns=columnlist_ID_removed)
added_risks.to_excel(writer,"New",index=True, columns=columnlist_ID_removed)

# counting no of closed risk and added since last review
now_closed = closed_risks.shape[0]
now_added = added_risks.shape[0]

# Adjusting column width
# get column values from output file into colwidth={} dictionary
for key, val in csv.reader(open("columnwidth.csv")):
    colwidth[key] = val

# counting changes for actions and rating
complete_actions=[]
total_complete_actions=[]
added_actions=[]
total_added_actions=[]
abandon_actions=[]
total_abandon_actions=[]
made_active = []
total_made_active = []
final=[]
total_rating = []
total_final=[]
total_now_added_closed = []
rating=[]
rating_rowno = None
action_rowno = None
no_rating_reduced = 0
no_rating_increased = 0
dates_updated =[]
dates_removed = []
dates_added =[]
due_dates =[]
cell_old_sorted = []
cell_new_sorted = []


# apply format to new sheets and get 'Actions Tracker'
for shit in book.worksheets:
    if shit == book.worksheets[0]:
        if changed_risks['Action Status'].str.contains("<<>>").any() or changed_risks['Current Risk Level'].str.contains("<<>>").any():
            # setting cell changes width and unhide if hidden
            shit.column_dimensions[get_column_letter(len(columnlist_ori)+1)].width = 38
            shit.column_dimensions[get_column_letter(len(columnlist_ori)+1)].hidden = False

    else:
        for col_letters in colwidth:
            shit.column_dimensions[col_letters].width = colwidth[col_letters]

        for rowNo in range(1, shit.max_row + 1):
            for columnNo in range(1, shit.max_column + 1):
                # setting the format. all string. all bordered and aligned
                cell = shit.cell(row=rowNo, column=columnNo)
                cell.value = str(cell.value)
                cell.alignment = alignment
                cell.border = border
                
                # colour top rows
                # if rowNo == 1:
                if cell.value in columnlist:
                    cell.fill = greyFill
                
                # 'Actions Tracker'
                elif shit == book.worksheets[1] and cell.value.find("<<>>") != -1 and cell.value is not None:
                    # define variables
                    cell.fill=redFill
                    cell_changed = shit.cell (row=rowNo, column=shit.max_column)
                    cell_ID = shit.cell (row=rowNo, column=shit.min_column) # column will always be minimum as Index will always be leftmost...
                    
                    # getting the comment cell row no in 'Comments' sheet. get index then plus excel firstrow value
                    comment_row = new[new['ID'].isin([cell_ID.value])].index.tolist()
                    comment_row = comment_row[0] + firstrow_excel # offset the index with first row from excel
                    cell_comment = book.worksheets[0].cell(row=comment_row, column=book.worksheets[0].max_column-1) # re-define 'Comments' worksheet and comment cell

                    # risk rating changes
                    if columnNo == columnlist.index('Current Risk Level')+1:
                        rating_old, rating_new = cell.value.split('\n--<<>>--\n')
                        rating_oldsplit = rating_old.split(' : ')
                        rating_newsplit = rating_new.split(' : ')
                        if int(rating_newsplit[0]) < int(rating_oldsplit[0]):
                            
                            # rating reduced
                            rating = 'Rating reduced from ' + rating_old + ' --> ' + rating_new
                            rating_rowno = rowNo
                            no_rating_reduced += 1
                            cell_comment.value = rating
                            cell_changed.value = rating
                            cell_comment.fill = greenFill
                            cell_comment.alignment = alignment
                            cell_comment.border = border
                        else:
                            # rating increased
                            rating = 'Rating increased from ' + rating_old + ' --> ' + rating_new
                            rating_rowno = rowNo
                            no_rating_increased += 1
                            cell_comment.value = rating
                            cell_changed.value = rating
                            cell_comment.fill = redFill
                            cell_comment.alignment = alignment
                            cell_comment.border = border
                    
                    # column 'Action Status' to track changes
                    elif columnNo == columnlist.index('Action Status')+1:
                        
                        # call compare & sort function to sort the changed cell
                        status_both = compare_sort (cell.value)
                        
                        for a,b in status_both: # a,b is the item in each tuple. a is old, b is new
                            if a != b: # and (x[0] is not None or x[0] != 'nan'):
                                #D = '{} --> {}'.format(a,b)

                                if a != None and a != ' - ': # None is from the shorter list above
                                    actionID = re.findall(r'\b\d+\b', b) # get actionID using regex

                                    # completed actions
                                    if b.endswith('Complete'):
                                        complete_actions.extend(actionID) # add corresponding actionID to completed_actions list
                                        total_complete_actions.extend(actionID) # add corresponding actionID to total completed_actions list to count total

                                    # actions made active
                                    if b.endswith ('Active'):
                                        made_active.extend(actionID) # add corresponding actionID to made_active list
                                        total_made_active.extend(actionID) # add corresponding actionID to total made_active list to count total

                                    # abandoned actions
                                    if b.endswith('Abandon'):
                                        abandon_actions.extend(actionID) # add corresponding actionID to abandon_actions list
                                        total_abandon_actions.extend(actionID) # add corresponding actionID to total abandon_actions list to count total
                                    
                                    # # new action
                                    # if a == ' - ':
                                    #     print ('new action')
                                    #     # actionID = re.findall(r'\b\d+\b', b) # get actionID using regex
                                    #     added_actions.extend(actionID) # add corresponding actionID to added_actions list
                                    #     total_added_actions.extend(actionID) # add corresponding actionID to total added_actions list to count total                                    

                                else: # new actions added
                                    actionID = re.findall(r'\b\d+\b', b) # get actionID using regex
                                    added_actions.extend(actionID) # add corresponding actionID to added_actions list
                                    total_added_actions.extend(actionID) # add corresponding actionID to total added_actions list to count total

                        # join all actionIDs into 1 string. this is so we can display them in 1 line
                        all_complete_actionsID = ', '.join(complete_actions)
                        all_added_actionsID = ', '.join(added_actions)
                        all_made_activeID = ', '.join(made_active)
                        all_abandon_actionsID = ', '.join(abandon_actions)

                        # display only if there is any value then append to the final list
                        if len(complete_actions) != 0:
                            final.append (str(len(complete_actions)) + ' Action(s) Completed (' + all_complete_actionsID + ')')

                        if len(added_actions) != 0:
                            final.append (str(len(added_actions)) + ' New Action(s) Added (' + all_added_actionsID + ')')
                        
                        if len(made_active) != 0:
                            final.append (str(len(made_active)) + ' Action(s) Made Active (' + all_made_activeID + ')')

                        if len(abandon_actions) != 0:
                            final.append (str(len(abandon_actions)) + ' Action(s) Abandoned (' + all_abandon_actionsID + ')')

                        # final list to 1 string and add to the cell
                        action_rowno = rowNo
                        if rating_rowno == action_rowno:
                            cell_comment.value = rating + '\n\n' + '\n'.join(final)
                            cell_changed.value = rating + '\n\n' + '\n'.join(final)
                        else:
                            cell_comment.value = '\n'.join(final)
                            cell_changed.value = '\n'.join(final)
                            cell_comment.fill = greenFill
                            cell_comment.alignment = alignment
                            cell_comment.border = border                            
                        #print('action ' + str(rowNo))

                        # clear all list if not append will continue adding together with previous values
                        del complete_actions[:]
                        del abandon_actions[:]
                        del added_actions[:]
                        del made_active[:]
                        del final[:]
                        del cell_old_sorted [:]
                        del cell_new_sorted [:]
                    
                    # column 'Due Dates' to track changes
                    elif columnNo == columnlist.index('Due Date')+1:
                        
                        # call compare & sort function to sort the changed cell
                        due_both = compare_sort (cell.value)
                        
                        for c,d in due_both: # c,d is the item in each tuple. c is old, d is new
                            if c != d: # means there is a change
                                if c != None: # means the action date is moved or added. if None, the action is new no need to track
                                    # split into list
                                    c = splitdash(c)
                                    d = splitdash(d)

                                    # c[1] is the due date, c[0] is the action ID. same goes for d
                                    # dates moved or updated
                                    if c[1] != '' and d[1] != '' and datetime.strptime(c[1], '%d %b %y').date() < datetime.strptime(d[1], '%d %b %y').date():                                    
                                        dates_updated.append(c[0]) # append action ID
                                    
                                    elif c[1] != '' and d[1] == '': # previously the action has a date but now removed
                                        dates_removed.append(c[0]) # append action ID

                                    elif c[1] == '': # previously the action has no date but now added
                                        dates_added.append(d[0]) # append action ID
                        
                        # join all action IDs into 1 string. this is so we can display them in 1 line string
                        all_dates_updated = ', '.join(dates_updated)
                        all_dates_removed = ', '.join(dates_removed)
                        all_dates_added = ', '.join(dates_added)
                        
                        # if dates_updated, dates_added, or dates_removed not empty, append to final due_dates list
                        if dates_updated:
                            due_dates.append (str(len(dates_updated)) + ' Date(s) Updated (' + all_dates_updated + ')')
                        
                        if dates_removed:
                            due_dates.append (str(len(dates_removed)) + ' Date(s) Removed (' + all_dates_removed + ')')

                        if dates_added:
                            due_dates.append (str(len(dates_added)) + ' Date(s) Added (' + all_dates_added + ')')

                        # write to cell.changed
                        if cell_changed.value != '': # easier this way. if there's value, we'll just add our comment. if not, write new one.
                        	cell_changed.value += '\n\n' + '\n'.join(due_dates)
                        else:
                        	cell_changed.value = '\n'.join(due_dates)

                        # clear all lists
                        del dates_added [:]
                        del dates_updated [:]
                        del dates_removed [:]
                        del due_dates [:]
                        del cell_old_sorted [:]
                        del cell_new_sorted [:]

# display total new and/or closed risk only if there is any value then append to total added closed list
if now_added != 0:
    total_now_added_closed.append(str(now_added) +' New Risk(s) Added')
if now_closed != 0:
    total_now_added_closed.append(str(now_closed) + ' Risk(s) Now Closed')

# display total risk rating changes only if there is any value then append to total rating list
if no_rating_reduced != 0:
    total_rating.append(str(no_rating_reduced) +' Risk Rating(s) Reduced')
if no_rating_increased != 0:
    total_rating.append(str(no_rating_increased) +' Risk Rating(s) Increased')

# display total actions changes only if there is any value then append to total final list
if len(total_complete_actions) != 0:
    total_final.append(str(len(total_complete_actions)) +' Action(s) Completed')

if len(total_added_actions) != 0:
    total_final.append(str(len(total_added_actions)) + ' New Action(s) Added')

if len(total_made_active) != 0:
    total_final.append(str(len(total_made_active)) + ' Action(s) Made Active')

if len(total_abandon_actions) != 0:
    total_final.append(str(len(total_abandon_actions)) + ' Action(s) Abandoned')

# rating, final action, and added_closed list into string
total_now_added_closed = '\n'.join(total_now_added_closed)
if total_now_added_closed != '':
    total_now_added_closed = total_now_added_closed + '\n'

total_rating = '\n'.join(total_rating)
if total_rating != '':
    total_rating = total_rating + '\n\n'

total_final = '\n'.join(total_final)
if total_final != '':
    total_final = total_final + '\n'

# re-define 'sheet' as wb has changed. sheet named 'Comments'. This is new sheet with comments. Previously max_column different because wb is original sheet with no extra comments.
sheet = book.get_sheet_by_name('Comments')
cell_final = sheet.cell(row=firstrow_excel, column=sheet.max_column)

# update final count
FINAL_SUMMARY = FINAL_SUMMARY + '\n\nChanges from previous risk review:\n' + total_now_added_closed + total_rating + total_final
cell_final.value = FINAL_SUMMARY

# new risk cell highlight
for rows in range(firstrow_excel, sheet.max_row + 1):
    cell = sheet.cell(row=rows, column=COLUMN_NO['ID'])
    if int(cell.value) in added_risks_set:
      cell.fill = greenFill

writer.save()

print('Done!\n\nExiting...\n\n\nDeveloped by Hazman Yusoff')
time.sleep(2)
