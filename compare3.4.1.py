#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  comp2.py
#  
#  Copyright 2017 yus22159 <yus22159@RP0005577591>
#  
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#  
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#  
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#  
#  

import pandas as pd
import collections
import csv
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side

# The previous method of using the dropna gives error when both files contain same nan column. this method does not use dropna but select the columns we want instead when saving.

# Define the diff function to show the changes in each field
def report_diff(x):
    if x[0] != x[1]:# and (x[0] is not None or x[0] != 'nan'):
        return ('{}\n--<<>>--\n{}'.format(*x))
    else:
        return x[0]
	#return x[0] if x[0] == x[1] else '{}\n--<<>>--\n{}'.format(*x)

alignment = Alignment(horizontal='general',
                      vertical='top',
                      text_rotation=0,
                      wrap_text=True,
                      shrink_to_fit=False,
                      indent=0)

border = Border(left=Side(border_style='thin',
                          color='B2B2B2'),
                right=Side(border_style='thin',
                           color='B2B2B2'),
                top=Side(border_style='thin',
                         color='B2B2B2'),
                bottom=Side(border_style='thin',
                            color='B2B2B2'))

redFill = PatternFill(start_color='FFC7CE',
                      end_color='FFC7CE',
                      fill_type='solid')

greyFill = PatternFill(start_color='DCDCDC',
                      end_color='DCDCDC',
                      fill_type='solid')

# disable SettingWithCopyWarning 
pd.options.mode.chained_assignment = None  # default='warn'

#getting the first row and all column name
#read excel with no header to get accurate row number and fill nan with 'N/A' for easy tracking
old = pd.read_excel('03.xlsx',header=None).fillna('N/A')
#iterate rows into tuples then into list
cycle=list(old.itertuples(index=False))
#empty list for our column name
columnlist=[]

firstrow=0
while firstrow < len(cycle):
  for colu in cycle[firstrow]:
    columnlist.append(colu)
  counter=collections.Counter(columnlist)
  if counter['N/A'] < 5:
    break
  firstrow+=1
  columnlist=[]

#final column name
columnlist = [x for x in columnlist if x != 'N/A']

# Read in the two files but call the data old and new and create columns to track
old = pd.read_excel('old.xlsx', header=firstrow)#.dropna(axis=1,how='all') if i dont dropna now, columns positions are still in placed
new = pd.read_excel('03.xlsx', header=firstrow)#.dropna(axis=1,how='all') 
old['version'] = "old"
new['version'] = "new"

# Set the columns you want to check to remove duplicate

# Set the order of the columns that'll be sorted when saving later
# Since the columns order is automatically retrieved, the lists are not needed

# empty list to get column values
colwidth = {}

#Join all the data together and ignore indexes so it all gets added
full_set = pd.concat([old,new],ignore_index=True).fillna('') #dropna here and columns order is still intact. not using anymore...

# Let's see what changes in the main columns we care about - subset is crucial if not it won't remove duplicates'
changes = full_set.drop_duplicates(subset=columnlist,keep='last')

#We want to know where the duplicate RiskIDs are, that means there have been changes. this return only the duplicated RiskID
dupe_risks = changes.set_index("ID").index.get_duplicates()

#Get all the duplicate rows. Subset dataframe
dupes = changes[changes["ID"].isin(dupe_risks)]

#Pull out the old and new data into separate dataframes
change_new = dupes[(dupes["version"] == "new")]
change_old = dupes[(dupes["version"] == "old")]

#Drop the temp columns - we don't need them now
change_new = change_new.drop(['version'], axis=1)
change_old = change_old.drop(['version'], axis=1)

#Index on the RiskID
change_new.set_index('ID',inplace=True)
change_old.set_index('ID',inplace=True)

#Now we can diff because we have two data sets of the same size with the same index
diff_panel = pd.Panel(dict(df1=change_old,df2=change_new))
changed_risks = diff_panel.apply(report_diff, axis=0)

#Diff'ing is done, we need to get a list of removed items

#Flag all duplicated RiskIDs
changes['duplicate']=changes["ID"].isin(dupe_risks)

#Identify non-duplicated items that are in the old version and did not show in the new version
closed_risks = changes[(changes["duplicate"] == False) & (changes["version"] == "old")]

#Re-Index Closed Risks
closed_risks.set_index('ID',inplace=True)

# We have the old and diff, we need to figure out which ones are new

#Drop duplicates but keep the first item instead of the last
new_risk_set = full_set.drop_duplicates(subset=columnlist,keep='first')

#Identify dupes in this new dataframe
new_risk_set['duplicate']=new_risk_set["ID"].isin(dupe_risks)

#Identify added risks
added_risks = new_risk_set[(new_risk_set["duplicate"] == False) & (new_risk_set["version"] == "new")]

#Re-Index added risks
added_risks.set_index('ID',inplace=True)

#Save the changes to excel but only include the columns we care about
book=load_workbook("03-comments.xlsx")
writer = pd.ExcelWriter("03-comments.xlsx",engine='openpyxl')
writer.book=book

#write on top of active sheet - no overwrite - no adding new sheets
writer.sheets=dict((ws.title,ws) for ws in book.worksheets)

#print(changed_risks.index.name) #to remove the index column from the columnlist
columnlist.remove(changed_risks.index.name) 
#save to an existing file in a new sheet and select the columns we want
changed_risks.to_excel(writer,"Changed", index=True, columns=columnlist)
closed_risks.to_excel(writer,"Closed",index=True, columns=columnlist)
added_risks.to_excel(writer,"New",index=True, columns=columnlist)

#adjusting column width
sheetchanged = book.get_sheet_by_name('Changed')
sheetclosed = book.get_sheet_by_name('Closed')
sheetnew = book.get_sheet_by_name('New')

#get column values from output file into colwidth={} dictionary
for key, val in csv.reader(open("columnwidth.csv")):
    colwidth[key] = val

#apply column values to new sheets
for col_letters in colwidth:
  sheetchanged.column_dimensions[col_letters].width = colwidth[col_letters]
  sheetclosed.column_dimensions[col_letters].width = colwidth[col_letters]
  sheetnew.column_dimensions[col_letters].width = colwidth[col_letters]

#apply format to new sheets
for shet in book.worksheets:
  if shet == book.worksheets[0]:
    continue
  else:
    for rowNo in range(1, shet.max_row + 1):
      for columnNo in range(1, shet.max_column + 1):
        cell = shet.cell(row=rowNo, column=columnNo)
        cell.value = str(cell.value)
        cell.alignment = alignment
        cell.border = border
        if rowNo==1:
          cell.fill = greyFill
        #elif cell.value=="nan\n--<<>>--\nnan" or cell.value=="NaT\n--<<>>--\nNaT":
          #cell.value=None
        elif cell.value.find("<<>>") != -1 and cell.value is not None:
          cell.fill=redFill

writer.save()
