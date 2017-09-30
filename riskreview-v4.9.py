# auto review risk register
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, Side
import time
from datetime import datetime, date
from tkinter.filedialog import askopenfilename

today = date.today()  # getting the date today.
# todaysdate=time.strftime("%d %b %y")
# print(todaysdate)

filename=askopenfilename()
if filename is '':
	print('No file selected')
	quit()

wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('03. RAPID Risk Register')

redFill = PatternFill(start_color='FFC7CE',
                      end_color='FFC7CE',
                      fill_type='solid')
yellowFill = PatternFill(start_color='FFFFCC', end_color='FFFFCC',
                         fill_type='solid')
italic = Font(size=10, italic=True)
alignment = Alignment(horizontal='general',
                      vertical='top',
                      text_rotation=0,
                      wrap_text=True,
                      shrink_to_fit=False,
                      indent=0)
border = Border(left=Side(border_style='thin',
                          color='B2B2B2'),
                right=Side(border_style='thin',
                           color='B2B2B2'),
                top=Side(border_style='thin',
                         color='B2B2B2'),
                bottom=Side(border_style='thin',
                            color='B2B2B2'))

def splitline (abc):
	return abc.split('\n')

def splitdash (xyz):
	return xyz.split(' - ')

# comments row
sheet.row_dimensions[5].height = 100
sheet['B5'] = 'Comments'
sheet['D5'] = 'Refer to ARM App User guide in step 2 page 19 and update where required'
sheet['E5'] = 'Refer to ARM App User guide in step 2 page 20 and update where required'
sheet['G5'] = 'Refer to ARM App User guide in step 2 page 21 - 23 and update where required.\n\nPlease review the unapproved risks'
sheet['H5'] = 'Refer to ARM App User guide in step 2 page 24 and update where required'
sheet['I5'] = 'Refer to ARM App User guide in step 2 page 25 and update where required'
sheet['K5'] = 'Refer to ARM App User guide in step 2 page 26 - 27 and update where required'
sheet['L5'] = 'Refer to ARM App User guide in step 2 page 28 - 29 and update where required'
sheet['M5'] = 'Refer to ARM App User guide in step 2 page 30 and update'
sheet['N5'] = 'Refer to ARM App User guide in step 2 page 32-33 and update where required'
sheet['O5'] = 'Refer to ARM App User guide in step 2 page 34 and update where required'
sheet['U5'] = 'Refer to ARM App User guide in step 2 page 62 and update.\nReview and update providing a brief concise note of how you arrived at the assessment\n"Probability of X% based on …"\n"Impact of X-Y days to recover"'
sheet['V5'] = 'Refer to ARM App User guide in step 2 page 49 and update.\n\nNote the dates past the resolution date. Have these been resolved? If so has the risk been updated or if not what is happening?'
sheet['W5'] = 'Refer to ARM App User guide in step 2 page 50 and update where required.\n\nPlease  note these dates should be aligned to your schedule activity that you have noted in the activity field where applicable'
sheet['X5'] = 'Refer to ARM App User guide in step 2 page 51 and update where required.\n\nPlease  note these dates should be aligned to your schedule activity that you have noted in the activity field where applicable'
sheet['AD5'] = 'Any high risk to be reviewed and approved?'
sheet['AK5'] = 'Refer to ARM App User guide in step 2 page 79-80 and update where required.\n\nAre there any additional actions?'
sheet['AL5'] = 'Refer to ARM App User guide in step 2 page 81 and update where required.\n\nHave the action owners changed for any outstanding actions?'
sheet['AM5'] = 'Refer to ARM App User guide in step 2 page 82 and update where required.\n\nHave any active actions been completed?'
sheet['AN5'] = 'Refer to ARM App User guide in step 2 page 83 and update where required.'
sheet['AR5'] = 'Are there any overdue actions?\nExclamation marks (!!!) at the end means the action is overdue'
sheet['AS5'] = 'Have all completed actions been updated and if completed then has the current risk rating been re-assessed?'

# list to compare with a cell
UPDATE = ['Undefined', 'Unapproved', '0 : NIL']
OVERDUE = []
COLUMN = {'Folder': int,
          'ID': int,
          'Title': int,
          'Owner': int,
          'Status': int,
          'Cause': int,
          'Effect': int,
          'Current Controls': int,
          'Description': int,
          'WBS': int,
          'Key Element': int,
          'Category': int,
          'Reviewed By': int,
          'Last Review': int,
          'Next Review Date': int,
          'Last Review Note': int,
          'Rationale': int,
          'Target Resolution': int,
          'Trigger Date': int,
          'Expiry Date': int,
          'Current Probability (Qualitative)': int,
          'Current Schedule Impact': int,
          'Current Cost Impact': int,
          'Current Reputation Impact': int,
          'Current HSE Impact': int,
          'Current Risk Level': int,
          'Target Probability (Qualitative)': int,
          'Target Schedule Impact': int,
          'Target Cost Impact': int,
          'Target Reputation Impact': int,
          'Target HSE Impact': int,
          'Target Risk Level': int,
          'Action Title': int,
          'Action Owner': int,
          'Action Status': int,
          'Action Type': int,
          'Action Description': int,
          'Action Comments': int,
          'Action Start Date': int,
          'Due Date': int,
          'Completion Date': int}

# looping the sheet
for rowNum in range(5, sheet.max_row + 1):
    for columnNum in range(1, sheet.max_column + 1):

        # setting up variables for loop
        cell = sheet.cell(row=rowNum, column=columnNum)

        # settign the conditions
        if rowNum == 5 and cell.value is not None:
            cell.font = italic
            cell.fill = yellowFill
            cell.alignment = alignment
            cell.border = border

        # getting column number into variable
        elif rowNum == 6 and cell.value in COLUMN:
            COLUMN[cell.value] = columnNum

        # cell is blank or in UPDATE
        elif rowNum > 6 and ((cell.value == '' or cell.value is None or cell.value in UPDATE) or 
                          (columnNum == COLUMN['Title'] and cell.value.startswith("Delay")) or
                          ((columnNum == COLUMN['Current Cost Impact']
                                  or columnNum == COLUMN['Current Reputation Impact']
                                  or columnNum == COLUMN['Current HSE Impact']
                                  or columnNum == COLUMN['Target Cost Impact']
                                  or columnNum == COLUMN['Target Reputation Impact']
                                  or columnNum == COLUMN['Target HSE Impact']) and cell.value != 'NIL') or
                          (columnNum == COLUMN['Action Owner'] and cell.value.find("Undefined") != -1)
                          ):
            cell.fill = redFill

		# current controls and action title	cross reference
        elif rowNum > 6 and (columnNum == COLUMN['Current Controls'] or columnNum == COLUMN['Action Title']):
            cxreflist = splitline(cell.value)
            cxreflist = [x for x in cxreflist if x != ''] # deleting blank lines
            for idex in cxreflist:
                #if idex.strip().find("(C") == -1 and idex.strip().find("[C") == -1 and idex.strip().find("[c") == -1 and idex.strip().find("(c") == -1 and idex.strip().find("- C") ==-1 and idex.strip().find("( C") ==-1:
                action = idex.replace(" ", "").lower()
                if action.find("(c") == -1 and action.find("[c") == -1:# and idex.strip().find("[c") == -1 and idex.strip().find("(c") == -1 and idex.strip().find("- C") ==-1 and idex.strip().find("( C") ==-1:
                  cell.fill = redFill
                else:
                    continue
        
        # current rating lesser than target rating
        elif rowNum > 6 and columnNum == COLUMN['Current Risk Level']:
            tarrate = sheet.cell(row=rowNum, column=COLUMN['Target Risk Level'])
            curratelist = cell.value.split(' : ')  # split the numbers and word
            tarratelist = tarrate.value.split(' : ')  # split the numbers and word
            if int(curratelist[0]) < int(tarratelist[0]):  # compare the number rating
                cell.fill = redFill

        # actions
        elif rowNum > 6 and columnNum == COLUMN['Action Status']:
            currate = sheet.cell(row=rowNum, column=COLUMN['Current Risk Level'])
            # actions that are completed but rating not re-assessed
            if cell.value.find("Active") == -1 and currate.value != tarrate.value:  # compare current and target rating
                cell.fill = redFill

            # overdue actions
            else:
                #stsplt = cell.value.split('\n')  # getting each line of action status into list
                stslist = splitline(cell.value)
                duedates = sheet.cell(row=rowNum, column=COLUMN['Due Date'])

                for idx, val in enumerate(stslist):
                    if val.endswith("Active"):
                        duedatelist = splitline(duedates.value)#.split('\n')  # getting each line of action duedate into list
                        date_line = splitdash(duedatelist[idx])#.split(' - ')  # splitting the action id and date into list. d[0] is the action id. d[1] is the date to compare.
                        if date_line[1] != '':  # if the date is not empty
                            due = datetime.strptime(date_line[1], '%d %b %y').date()  # converting str type into datetime type
                            if due < today:  # compare with today's date
                                duedatelist[idx] = duedatelist[idx] + " !!!" #this will just update the corresponding date in list
                                OVERDUE.append(idx)
                                #duedates.value = newduedates
                                duedates.value = '\n'.join(duedatelist)
                                duedates.fill = redFill
                        else:  # if date is empty
                            duedates.fill = redFill

sheet['AR4'] = str(len(OVERDUE)) + ' Overdue Action(s)'
sheet.title='Comments'
wb.save('03-comments.xlsx')
